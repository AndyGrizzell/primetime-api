from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
import copy

app = Flask(__name__)
CORS(app)

GREEN = "FF00B050"
BLACK = "FF000000"

def green_fill():
    return PatternFill("solid", fgColor=GREEN)

def bold_font(size=10):
    return Font(name="Arial", bold=True, size=size)

def normal_font(size=10):
    return Font(name="Arial", bold=False, size=size)

def style_green_row(ws, row_num, max_col=6):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = green_fill()
        cell.font = bold_font()

def style_bold_row(ws, row_num, max_col=6):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = bold_font()

def fmt_currency(val):
    return val if val else 0

@app.route('/health', methods=['GET'])
def health():
    return jsonify({"status": "ok"})

@app.route('/generate-buildsheet', methods=['POST', 'OPTIONS'])
def generate_buildsheet():
    if request.method == 'OPTIONS':
        return '', 204

    d = request.json
    qty = int(d.get('quantity', 1) or 1)
    vins = [d.get(f'vin_{i}', '') or '' for i in range(1, qty + 1)]

    import os
    template_path = os.path.join(os.path.dirname(__file__), 'BS_Final.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # --- CUSTOMER INFO ---
    ws['B2'] = d.get('organizationName', '')
    ws['D2'] = str(qty)
    ws['B3'] = d.get('address', '')
    ws['D3'] = d.get('chassisRelease', '')
    ws['B4'] = d.get('cityState', '')
    ws['D4'] = d.get('finCode', '')
    ws['B5'] = d.get('contact', '')
    ws['B6'] = d.get('phone', '')
    ws['B7'] = d.get('email', '')
    ws['D8'] = d.get('date', '')
    ws['B9'] = vins[0] if vins else ''
    ws['E10'] = 'Yes' if d.get('adaptiveMobility') == 'Yes' else 'No'

    # --- CHASSIS (E15-E25) ---
    chassis = d.get('chassis', '')
    ws['E15'] = 1 if '2026 Low Roof' in chassis and 'Builders Prep' in chassis else 0
    ws['E16'] = 1 if '2026 Low Roof' in chassis and '12 Passenger' in chassis else 0
    ws['E17'] = 1 if '2026 Low Roof' in chassis and '15 Passenger' in chassis else 0
    ws['E18'] = 1 if '2026 Mid Roof' in chassis and '12 Passenger' in chassis else 0
    ws['E19'] = 1 if '2026 Mid Roof' in chassis and 'Builders Prep' in chassis else 0
    ws['E20'] = 0  # Promaster Window
    ws['E21'] = 1 if 'Promaster' in chassis and 'Cargo' in chassis else 0
    ws['E22'] = 1 if d.get('fullBodyPaintOEM') == 'Yes' else 0
    ws['E23'] = 1 if d.get('fullBodyPaintNonOEM') == 'Yes' else 0
    ws['E24'] = 0
    ws['E25'] = 0

    # --- INTERIOR (E28-E34) ---
    upfit = d.get('interiorUpfit', '').split(' (+')[0] if d.get('interiorUpfit') else ''
    ws['E28'] = 1 if 'Ford Transit' in upfit and 'Side Rear' not in upfit else 0
    ws['E29'] = 1 if 'Side Rear Lift' in upfit else 0
    ws['E30'] = 1 if 'Promaster Window' in upfit else 0
    ws['E31'] = 1 if 'Promaster LF' in upfit else 0
    ws['E32'] = 1 if d.get('rearStorageBarrier') == 'Yes' else 0
    ws['E33'] = 1 if d.get('storageWalkerMount') == 'Yes' else 0
    ws['E34'] = 1 if d.get('paSystem') == 'Yes' else 0

    # --- RUNNING BOARD (E36-E38) ---
    ws['E36'] = 1 if d.get('passengerRunningBoard') == 'Yes' else 0
    ws['E37'] = 1 if d.get('driverRunningBoard') == 'Yes' else 0
    ws['E38'] = 1 if d.get('rearMudFlaps') == 'Yes' else 0

    # --- A/C (E42-E45) ---
    ac = d.get('acHeat', '')
    ws['E42'] = 1 if 'Twin' in ac else 0
    ws['E43'] = 1 if 'Dual' in ac else 0
    ws['E44'] = 1 if not ac or 'OEM' in ac else 0
    ws['E45'] = 0

    # --- FLOORING (E47-E53) ---
    fk = d.get('flooring', '').split(' (+')[0] if d.get('flooring') else ''
    ws['E47'] = 1 if 'Wood Grain' in fk else 0
    ws['E48'] = 1 if 'Altro' in fk and 'Wood' not in fk else 0
    ws['E49'] = 1 if 'OEM Seat Package' in fk else 0
    ws['E50'] = 1 if 'Pareto' in fk and 'Ford' in fk else 0
    ws['E51'] = 1 if 'Pareto' in fk and 'Dodge' in fk else 0
    ws['E52'] = 0
    ws['E53'] = 0

    # --- SEATING (E55-E65) ---
    ws['E55'] = int(d.get('seatSingleGO') or 0)
    ws['E56'] = int(d.get('seatDoubleGO') or 0)
    ws['E57'] = 0
    ws['E58'] = 0
    ws['E59'] = int(d.get('seatDoubleFoldaway') or 0)
    ws['E60'] = int(d.get('seatSingleFoldaway') or 0)
    ws['E61'] = int(d.get('seatPareto') or 0)
    ws['E62'] = 0
    ws['E63'] = int(d.get('seatBeltExtQty') or 0)
    ws['E64'] = int(d.get('seatARACPerimeter') or 0)
    ws['E65'] = int(d.get('seatArmRests') or 0)

    # --- WC DOOR (E67) ---
    ws['E67'] = 1 if d.get('wcDoor') == 'Yes' else 0

    # --- WC LIFT (E69-E78) ---
    selected_lift = d.get('wcLift', '').split(' (+')[0] if d.get('wcLift') else ''
    lift_row_map = {
        'Braun Century 34x51 #800': 69,
        'Braun Century 34x51 #1000': 70,
        'Braun Century 37x54 #1000': 71,
        'Braun Century Rear Side Door 34x51 #1000': 72,
        'Braun Millenium 34x51 #800': 73,
        'Braun Millenium 34x51 #1000': 74,
        'Braun Shift N Step Lift': 76,
    }
    for r in [69,70,71,72,73,74,76]:
        ws[f'E{r}'] = 0
    if selected_lift in lift_row_map:
        ws[f'E{lift_row_map[selected_lift]}'] = 1
    ws['E77'] = 1 if d.get('adaInterlock') == 'Yes' else 0
    ws['E78'] = 1 if d.get('passengerCallBell') == 'Yes' else 0

    # --- WC RESTRAINTS (E80-E85) ---
    ws['E80'] = int(d.get('lTrackQty') or 0)
    ws['E81'] = int(d.get('shoulderAnchor') or 0)
    ws['E82'] = 0
    ws['E83'] = int(d.get('qStraintLTrack') or 0)
    ws['E84'] = int(d.get('slideNClick') or 0)
    ws['E85'] = int(d.get('qStraintSlide') or 0)

    # --- DESTINATION SIGNAGE (E90-E91) ---
    ws['E90'] = 1 if d.get('frontDestSign') == 'Yes' else 0
    ws['E91'] = 1 if d.get('sideDestSign') == 'Yes' else 0

    # --- STANTIONS (E94-E97) ---
    ws['E94'] = 1 if d.get('entranceGrabBar') == 'Standard (+$199)' else 0
    ws['E95'] = 1 if d.get('entranceGrabBar') == 'Yellow (+$189)' else 0
    ws['E96'] = 1 if d.get('parallelGrabBars') == 'Yes' else 0
    ws['E97'] = 1 if d.get('stantions') == 'Yes' else 0

    # --- ENTRANCE DOOR (E99-E104) ---
    ed = d.get('entranceDoor', '')
    ws['E99'] = 1 if 'Standard' in ed else 0
    ws['E100'] = 1 if 'L.F.' in ed else 0
    ws['E101'] = 1 if d.get('keyedRemoteEntry') == 'Yes' else 0
    ws['E102'] = 1 if d.get('remoteEntry') == 'Yes' else 0
    ws['E103'] = 0
    ws['E104'] = 0

    # --- SAFETY (E106-E111) ---
    ws['E106'] = 1 if d.get('safetyKit') == 'Yes' else 0
    ws['E107'] = 1 if d.get('roofHatch') == 'Yes' else 0
    strobe = d.get('strobeLight', '')
    ws['E108'] = 1 if 'Color' in strobe else 0
    ws['E109'] = 1 if d.get('heightDecal') == 'Yes' else 0
    ws['E110'] = 1 if d.get('watchStepDecal') == 'Yes' else 0
    ws['E111'] = 1 if 'Clear' in strobe else 0

    # --- AUDIO (E113-E116) ---
    ws['E113'] = 1 if d.get('paSystemAudio') == 'Yes' else 0
    ws['E114'] = 1 if d.get('externalSpeaker') == 'Yes' else 0
    ws['E115'] = 1 if d.get('lockableStorageWood') == 'Yes' else 0
    ws['E116'] = 1 if d.get('lockableStorageSteel') == 'Yes' else 0

    # --- ELECTRICAL (E118-E121) ---
    ws['E118'] = 1 if d.get('upgradedDomeLights') == 'Yes' else 0
    ws['E119'] = 0
    ws['E120'] = 1 if d.get('heatedStepWell') == 'Yes' else 0
    ws['E121'] = int(d.get('usbPorts') or 0)

    # --- SPECIAL BUILDS (E128-E134) ---
    sb = d.get('specialBuild', '')
    ws['E128'] = 1 if '10 Passenger' in sb else 0
    ws['E129'] = 1 if '15 Passenger' in sb else 0
    ws['E130'] = 1 if 'Seat Package B' in sb else 0
    ws['E131'] = 1 if d.get('lockableStorageBox') == 'Yes' else 0
    ws['E132'] = 1 if d.get('fairboxPrewire') == 'Yes' else 0

    # --- BSI SUPPLIED (E145-E155) ---
    ws['E145'] = 1 if d.get('basicGraphics') == 'Yes' else 0
    ws['E146'] = 0
    ws['E147'] = 1 if d.get('bsiAddOns') == 'Yes' else 0
    ws['E148'] = 1 if d.get('angeltrax') == 'Yes' else 0
    ws['E149'] = 1 if d.get('undercoat') == 'Yes' else 0
    ws['E150'] = 1 if d.get('customGraphics') == 'Yes' else 0
    ws['E151'] = 1 if d.get('oemSeatPackage') == 'Yes' else 0
    ws['E152'] = 1 if d.get('classHitch') == 'Yes' else 0
    ws['E153'] = 0
    ws['E154'] = 0
    ws['E155'] = 1 if d.get('schoolSign') == 'Yes' else 0

    # --- INCENTIVE (E158) ---
    ws['E158'] = 1 if d.get('mobilityIncentive') == 'Yes' else 0

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    org = (d.get('organizationName') or 'Build').replace(' ', '_')
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'BSI_{org}_BuildSheet.xlsx')



@app.route('/generate-proposal', methods=['POST', 'OPTIONS'])
def generate_proposal():
    if request.method == 'OPTIONS':
        return '', 204

    d = request.json
    qty = int(d.get('quantity', 1) or 1)
    vins = [d.get(f'vin_{i}', '') or '' for i in range(1, qty + 1)]

    pt = calculate_pt(d)
    bsi = calculate_bsi(d)
    grand = pt + bsi

    # Load original as template - preserves all formatting, colors, borders, images, merged cells
    import os
    template_path = os.path.join(os.path.dirname(__file__), '2026_PRIME_FORD_TRANSIT_PROPOSAL.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # --- CUSTOMER INFO ---
    ws['A2'] = d.get('organizationName', '')
    ws['C2'] = d.get('address', '')
    ws['C3'] = d.get('cityState', '')
    ws['H2'] = d.get('contact', '')
    ws['H3'] = d.get('phone', '')
    ws['H4'] = d.get('email', '')
    ws['B9'] = vins[0] if vins else ''
    ws['H10'] = d.get('salesperson', '')
    ws['B11'] = d.get('seatMaterial', 'Vinyl')
    ws['B10'] = 'Yes' if d.get('customGraphics') == 'Yes' else ''

    # --- CHASSIS (H42-H44) ---
    chassis = d.get('chassis', '')
    ws['H42'] = 0  # 136 Low Roof
    ws['H43'] = 1 if '12 Passenger' in chassis else 0  # 148 Mid Roof 12 pass
    ws['H44'] = 1 if '15 Passenger' in chassis else 0  # 148 Extended High Roof

    # Special build notes (H45-H46)
    notes = d.get('specialNotes', '')
    ws['H45'] = 1 if notes else 0
    ws['H46'] = 1 if notes else 0
    ws['A46'] = notes or None

    # --- FLOORING (H53-H56) ---
    fk = d.get('flooring', '').split(' (+')[0] if d.get('flooring') else ''
    ws['H53'] = 1 if 'Altro' in fk and 'Wood' not in fk else 0
    ws['H56'] = 1 if 'Wood' in fk else 0

    # --- EXTERIOR GRAPHICS (H59-H60) ---
    ws['H59'] = 1 if d.get('fullBodyPaintOEM') == 'Yes' or d.get('fullBodyPaintNonOEM') == 'Yes' else 0
    ws['H60'] = 1 if d.get('customGraphics') == 'Yes' else 0

    # --- RUNNING BOARDS (H70, H72) ---
    ws['H70'] = 1 if d.get('passengerRunningBoard') == 'Yes' else 0
    ws['H72'] = 1 if d.get('driverRunningBoard') == 'Yes' else 0

    # --- A/C (H74-H75) ---
    ac = d.get('acHeat', '')
    ws['H74'] = 1 if not ac or 'OEM' in ac else 0
    ws['H75'] = 1 if 'Twin' in ac else 0

    # --- USB (H80) ---
    ws['H80'] = 1 if d.get('usbPorts') and int(d.get('usbPorts', 0)) > 0 else 0

    # --- DESTINATION SIGNS (H82-H83) ---
    ws['H83'] = 1 if d.get('frontDestSign') == 'Yes' else 0

    # --- STROBE (H86) ---
    strobe = d.get('strobeLight', '')
    ws['H86'] = 1 if strobe and strobe != 'No' else 0

    # --- INTERIOR LIGHTS (H89) ---
    ws['H89'] = 1 if d.get('upgradedDomeLights') == 'Yes' else 0

    # --- AUDIO (H92-H94) ---
    ws['H92'] = 1 if d.get('paSystem') == 'Yes' or d.get('paSystemAudio') == 'Yes' else 0
    ws['H94'] = 1 if d.get('externalSpeaker') == 'Yes' else 0

    # --- DOORS (H96-H101) ---
    ed = d.get('entranceDoor', '')
    ws['H96'] = 1 if not ed or ed == 'None' else 0
    ws['H97'] = 1 if 'Bi Fold' in ed else 0
    ws['H98'] = 1 if d.get('remoteEntry') == 'Yes' else 0
    ws['H99'] = 1 if d.get('keyedRemoteEntry') == 'Yes' else 0
    ws['H101'] = 1 if d.get('roofHatch') == 'Yes' else 0

    # --- WC LIFTS (H106-H113) ---
    selected_lift = d.get('wcLift', '').split(' (+')[0] if d.get('wcLift') else ''
    lift_map = {
        'Braun Century 34x51 #800': 107,
        'Braun Century 34x51 #1000': 109,
        'Braun Century 37x54 #1000': 110,
        'Braun Century Rear Side Door 34x51 #1000': 110,
        'Braun Millenium 34x51 #800': 112,
        'Braun Millenium 34x51 #1000': 112,
        'Braun Shift N Step Lift': 113,
    }
    for row_num in [106,107,108,109,110,111,112,113]:
        ws[f'H{row_num}'] = 0
    if selected_lift in lift_map:
        ws[f'H{lift_map[selected_lift]}'] = 1

    # --- ADA INTERLOCK (H117) ---
    ws['H117'] = 1 if d.get('adaInterlock') == 'Yes' else 0

    # --- Q-STRAINT (H119-H120, H124) ---
    ws['H119'] = int(d.get('qStraintSlide') or 0)
    ws['H120'] = int(d.get('qStraintLTrack') or 0)
    ws['H124'] = int(d.get('lTrackQty') or 0)

    # --- SAFETY (H135-H139) ---
    safety = d.get('safetyKit') == 'Yes'
    ws['H135'] = 1 if safety else 0
    ws['H136'] = 1 if safety else 0
    ws['H137'] = 1 if safety else 0
    ws['H138'] = 1 if safety else 0

    # --- STANCHIONS (H152) ---
    ws['H152'] = 1 if d.get('stantions') == 'Yes' else 0

    # --- SEATING (H160-H165, H169) ---
    ws['H160'] = int(d.get('seatDoubleGO') or 0)
    ws['H161'] = int(d.get('seatSingleGO') or 0)
    ws['H162'] = int(d.get('seatDoubleFoldaway') or 0)
    ws['H163'] = int(d.get('seatSingleFoldaway') or 0)
    ws['H169'] = int(d.get('seatArmRests') or 0)

    # --- SEAT BELTS (H175) ---
    ws['H175'] = int(d.get('seatBeltExtQty') or 0)

    # --- TERMS / PRICING ---
    # Find pricing cells (around row 182-200)
    ws['J182'] = pt
    mobility_rebate = -6451 if d.get('mobilityIncentive') == 'Yes' else 0
    ws['J186'] = mobility_rebate
    ws['I188'] = pt + mobility_rebate
    ws['J196'] = grand
    ws['I198'] = qty
    ws['I200'] = grand * qty

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    org = (d.get('organizationName') or 'Build').replace(' ', '_')
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'BSI_{org}_Proposal.xlsx')



def calculate_pt(d):
    total = 0
    chassis_prices = {
        "2026 Low Roof Chassis  - Builders Prep": 54489,
        "2026 Low Roof Chassis  - 12 Passenger": 55318,
        "2026 Low Roof Chassis  - 15 Passenger": 56608,
        "2026 Mid Roof Chassis  - 12 Passenger": 56030,
        "2026 Mid Roof Chassis  - Builders Prep": 55201,
        "2025 Promaster - Cargo": 49283,
    }
    if d.get('chassis') in chassis_prices:
        total += chassis_prices[d['chassis']]
    if d.get('fullBodyPaintOEM') == 'Yes': total += 329
    if d.get('fullBodyPaintNonOEM') == 'Yes': total += 7995

    upfit_prices = {"Interior Upfit - Ford Transit": 3995, "Interior Upfit - Ford Transit - Side Rear Lift": 4995,
                    "Interior Upfit - Promaster Window": 7995, "Interior Upfit - Promaster LF": 11995}
    uk = d.get('interiorUpfit', '').split(' (+')[0] if d.get('interiorUpfit') else ''
    if uk in upfit_prices: total += upfit_prices[uk]

    if d.get('paSystem') == 'Yes': total += 475
    if d.get('storageWalkerMount') == 'Yes': total += 395
    if d.get('passengerRunningBoard') == 'Yes': total += 390
    if d.get('driverRunningBoard') == 'Yes': total += 5
    if d.get('rearMudFlaps') == 'Yes': total += 75

    ac = d.get('acHeat', '')
    if 'Twin' in ac: total += 1895
    elif 'Dual' in ac: total += 5595

    floor_prices = {"Plywood Subfloor with Wood Grain Flooring": 1395, "Pareto Floor - Ford": 4995,
                    "Pareto Floor - Dodge": 5995, "Modify Flooring - OEM Seat Package": 1295}
    fk = d.get('flooring', '').split(' (+')[0] if d.get('flooring') else ''
    if fk in floor_prices: total += floor_prices[fk]

    total += int(d.get('seatSingleGO') or 0) * 827
    total += int(d.get('seatDoubleGO') or 0) * 1695
    total += int(d.get('seatDoubleFoldaway') or 0) * 2075
    total += int(d.get('seatSingleFoldaway') or 0) * 1195
    total += int(d.get('seatPareto') or 0) * 200
    total += int(d.get('seatArmRests') or 0) * 60
    total += int(d.get('seatBeltExtQty') or 0) * 30

    if d.get('wcDoor') == 'Yes': total += 7995
    lift_prices = {"Braun Century 34x51 #800": 5819, "Braun Century 34x51 #1000": 5995,
                   "Braun Century 37x54 #1000": 7195, "Braun Century Rear Side Door 34x51 #1000": 6995,
                   "Braun Millenium 34x51 #800": 5995, "Braun Millenium 34x51 #1000": 6995, "Braun Shift N Step Lift": 8699}
    lk = d.get('wcLift', '').split(' (+')[0] if d.get('wcLift') else ''
    if lk in lift_prices: total += lift_prices[lk]
    if d.get('adaInterlock') == 'Yes': total += 695
    if d.get('passengerCallBell') == 'Yes': total += 495

    total += int(d.get('lTrackQty') or 0) * 150
    total += int(d.get('shoulderAnchor') or 0) * 249
    total += int(d.get('qStraintLTrack') or 0) * 595
    total += int(d.get('slideNClick') or 0) * 52
    total += int(d.get('qStraintSlide') or 0) * 723

    if d.get('frontDestSign') == 'Yes': total += 3495
    if d.get('sideDestSign') == 'Yes': total += 1494

    ed = d.get('entranceDoor', '')
    if 'Standard' in ed: total += 5295
    elif 'L.F.' in ed: total += 6995
    if d.get('remoteEntry') == 'Yes': total += 75
    if d.get('keyedRemoteEntry') == 'Yes': total += 95
    if d.get('entranceGrabBar') == 'Standard (+$199)': total += 199
    if d.get('entranceGrabBar') == 'Yellow (+$189)': total += 189
    if d.get('parallelGrabBars') == 'Yes': total += 195
    if d.get('stantions') == 'Yes': total += 495

    if d.get('safetyKit') == 'Yes': total += 395
    if d.get('roofHatch') == 'Yes': total += 695
    if d.get('strobeLight') and d.get('strobeLight') != 'No': total += 395
    if d.get('heightDecal') == 'Yes': total += 20
    if d.get('watchStepDecal') == 'Yes': total += 20

    if d.get('upgradedDomeLights') == 'Yes': total += 100
    if d.get('heatedStepWell') == 'Yes': total += 347.5
    total += int(d.get('usbPorts') or 0) * 50
    if d.get('paSystemAudio') == 'Yes': total += 395
    if d.get('externalSpeaker') == 'Yes': total += 50
    if d.get('lockableStorageWood') == 'Yes': total += 395
    if d.get('lockableStorageSteel') == 'Yes': total += 495

    sb = d.get('specialBuild', '')
    if '10 Passenger' in sb: total += 1295
    elif '15 Passenger' in sb: total += 3000
    elif 'Seat Package B' in sb: total += 5495
    if d.get('lockableStorageBox') == 'Yes': total += 395
    if d.get('fairboxPrewire') == 'Yes': total += 30
    total += float(d.get('specialNotesPrice') or 0)

    total += -500 + 475  # drop ship credit + freight
    return total


def calculate_bsi(d):
    total = 0
    if d.get('basicGraphics') == 'Yes': total += 325
    if d.get('bsiAddOns') == 'Yes': total += 350
    if d.get('angeltrax') == 'Yes': total += 2000
    if d.get('undercoat') == 'Yes': total += 400
    if d.get('oemSeatPackage') == 'Yes': total += 1200
    if d.get('classHitch') == 'Yes': total += 499
    if d.get('schoolSign') == 'Yes': total += 325
    return total


if __name__ == '__main__':
    app.run(debug=True, port=5000)
